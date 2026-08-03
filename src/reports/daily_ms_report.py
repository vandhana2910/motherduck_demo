"""
Market Share Dashboard Generator
================================
Reads directly from your MotherDuck (dbt_demo) fact tables:
    - fct_weekend_rollup       (one row per venue, Fri+Sat+Sun summed)
    - fct_wtd_rollup           (one row per venue, full fiscal week summed)
    - fct_wtd_daily_combined   (one row per venue per day)

Generates ONE Excel workbook (market_share_report.xlsx) with a SINGLE VISIBLE
"Dashboard" sheet:
    - A "Fiscal Week" dropdown (data validation) defaulting to P12W2 if
      present, otherwise the latest week.
    - KPI cards (Total Weekend Revenue, Total WTD Revenue, WTD Market Share,
      WTD Industry Gross) that update to match the selected week.
    - The full block table (Weekend / WTD / each calendar day x 7 metrics),
      one row per venue plus a TOTAL row, that updates to match the selected
      week.

Every other fiscal week's data lives on hidden helper sheets (Data,
BlockLabels, Lists) and the Dashboard cells are FORMULAS (SUMIFS / INDEX-
MATCH) that filter on the dropdown - so picking a different week from the
dropdown re-displays that week's numbers without re-running this script.

Run:
    pip install duckdb pandas openpyxl
    python market_share_report.py
"""

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = "market_share_report.xlsx"
DEFAULT_WEEK = "P12W2"   # shown on open if present, else falls back to latest week
MAX_DAY_BLOCKS = 7       # a fiscal week is Mon-Sun at most

# ---------------------------------------------------------------------------
# Connect to MotherDuck
# ---------------------------------------------------------------------------
con = duckdb.connect()
con.execute("ATTACH 'md:dbt_demo'")  # will prompt browser login if no token set
con.execute("USE dbt_demo")

# ---------------------------------------------------------------------------
# Column mappings (same generic METRICS across all three tables)
# ---------------------------------------------------------------------------
METRICS = ["Gross", "Industry", "Share", "PY Gross", "PY Industry", "PY", "Delta"]

WEEKEND_COLS = {
    "Gross": "weekend_gross", "Industry": "weekend_industry", "Share": "weekend_share",
    "PY Gross": "py_weekend_gross", "PY Industry": "Py_Industry", "PY": "py_weekend_share",
    "Delta": "weekend_delta",
}
WTD_COLS = {
    "Gross": "wtd_gross", "Industry": "wtd_industry", "Share": "wtd_share",
    "PY Gross": "py_wtd_gross", "PY Industry": "py_wtd_industry", "PY": "py_wtd_share",
    "Delta": "wtd_delta",
}
DAILY_COLS = {
    "Gross": "Gross", "Industry": "Industry", "Share": "share",
    "PY Gross": "Py_Gross", "PY Industry": "Py_Industry", "PY": "Py_share",
    "Delta": "delta",
}

SUM_METRICS = ["Gross", "PY Gross"]
CARRY_METRICS = ["Industry", "PY Industry"]

CURRENCY_METRICS = {"Gross", "Industry", "PY Gross", "PY Industry"}
PERCENT_METRICS = {"Share", "PY", "Delta"}

CURRENCY_FMT = '$#,##0;($#,##0);"-"'
# Values already come out of SQL on a 0-100 scale, so append a literal "%"
# rather than using Excel's native 0.0% (which would divide by 100 again).
PERCENT_FMT = '0.0"%";(0.0"%");"-"'

FONT_NAME = "Arial"

BLOCK_KEYS = ["Weekend", "WTD"] + [f"Day{i}" for i in range(1, MAX_DAY_BLOCKS + 1)]
TOTAL_VENUE = "TOTAL"


# ---------------------------------------------------------------------------
# Data access
# ---------------------------------------------------------------------------
def get_available_weeks():
    df = con.execute("""
        SELECT fiscal_week, MIN(day_count) AS first_day
        FROM fct_wtd_daily_combined
        GROUP BY fiscal_week
        ORDER BY first_day
    """).df()
    return df["fiscal_week"].tolist()


def fetch_week_frames(week):
    weekend_df = con.execute(
        "SELECT * FROM fct_weekend_rollup WHERE fiscal_week = ?", [week]
    ).df()
    wtd_df = con.execute(
        "SELECT * FROM fct_wtd_rollup WHERE fiscal_week = ?", [week]
    ).df()
    daily_df = con.execute(
        "SELECT * FROM fct_wtd_daily_combined WHERE fiscal_week = ?", [week]
    ).df()
    return weekend_df, wtd_df, daily_df


# ---------------------------------------------------------------------------
# Shaping (same logic as before, blocks now keyed positionally: Weekend, WTD,
# Day1..Day7, so every week lines up under the same fixed column layout)
# ---------------------------------------------------------------------------
def build_blocks(weekend_df, wtd_df, daily_df):
    """Returns (day_labels, data, venues). day_labels maps DayN -> 'dd Mon yyyy'."""
    all_venues = set()
    for df in (weekend_df, wtd_df, daily_df):
        if "venue" in df.columns:
            all_venues.update(df["venue"].dropna().unique().tolist())
    venues = sorted(all_venues)

    dates = sorted(daily_df["date"].unique()) if not daily_df.empty else []
    day_labels = {}
    for i, d in enumerate(dates[:MAX_DAY_BLOCKS]):
        day_labels[f"Day{i + 1}"] = pd.Timestamp(d).strftime("%d %b %Y")

    data = {v: {b: {m: None for m in METRICS} for b in BLOCK_KEYS} for v in venues}

    def fill(block_key, row, colmap):
        v = row["venue"]
        if v not in data:
            return
        data[v][block_key] = {m: row[colmap[m]] for m in METRICS}

    for _, row in weekend_df.iterrows():
        fill("Weekend", row, WEEKEND_COLS)
    for _, row in wtd_df.iterrows():
        fill("WTD", row, WTD_COLS)

    for i, d in enumerate(dates[:MAX_DAY_BLOCKS]):
        block_key = f"Day{i + 1}"
        day_rows = daily_df[daily_df["date"] == d]
        for _, row in day_rows.iterrows():
            fill(block_key, row, DAILY_COLS)

    return day_labels, data, venues


def build_totals(data, venues):
    """Cumulative TOTAL row per block, summed across every venue."""
    totals = {b: {m: None for m in METRICS} for b in BLOCK_KEYS}
    for b in BLOCK_KEYS:
        row = {m: 0.0 for m in SUM_METRICS}
        industry = None
        py_industry = None
        any_data = False
        for v in venues:
            cell = data[v][b]
            if any(pd.notna(cell.get(m)) for m in METRICS):
                any_data = True
            for m in SUM_METRICS:
                val = cell.get(m)
                if pd.notna(val):
                    row[m] += val
            if industry is None and pd.notna(cell.get("Industry")):
                industry = cell.get("Industry")
            if py_industry is None and pd.notna(cell.get("PY Industry")):
                py_industry = cell.get("PY Industry")

        if not any_data:
            continue

        share = (row["Gross"] / industry * 100) if industry else None
        py_share = (row["PY Gross"] / py_industry * 100) if py_industry else None
        delta = (share - py_share) if (share is not None and py_share is not None) else None

        totals[b] = {
            "Gross": row["Gross"],
            "Industry": industry,
            "Share": share,
            "PY Gross": row["PY Gross"],
            "PY Industry": py_industry,
            "PY": py_share,
            "Delta": delta,
        }
    return totals


# ---------------------------------------------------------------------------
# Build long-format rows for the hidden "Data" and "BlockLabels" sheets
# ---------------------------------------------------------------------------
def collect_all_weeks():
    weeks = get_available_weeks()
    data_rows = []        # (FiscalWeek, Venue, BlockKey, Metric, Value)
    block_label_rows = []  # (FiscalWeek, BlockKey, DisplayLabel, HelperKey)
    all_venues = set()

    for week in weeks:
        weekend_df, wtd_df, daily_df = fetch_week_frames(week)
        day_labels, data, venues = build_blocks(weekend_df, wtd_df, daily_df)
        totals = build_totals(data, venues)
        all_venues.update(venues)

        for v in venues:
            for b in BLOCK_KEYS:
                for m in METRICS:
                    val = data[v][b][m]
                    if pd.notna(val):
                        data_rows.append((week, v, b, m, float(val)))

        for b in BLOCK_KEYS:
            trow = totals[b]
            for m in METRICS:
                val = trow.get(m)
                if val is not None and pd.notna(val):
                    data_rows.append((week, TOTAL_VENUE, b, m, float(val)))

        for block_key, label in day_labels.items():
            block_label_rows.append((week, block_key, label, f"{week}|{block_key}"))

    return weeks, sorted(all_venues), data_rows, block_label_rows


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------
def _fmt_for(metric):
    if metric in CURRENCY_METRICS:
        return CURRENCY_FMT
    if metric in PERCENT_METRICS:
        return PERCENT_FMT
    return "General"


def write_hidden_sheets(wb, weeks, data_rows, block_label_rows):
    # --- Data sheet ------------------------------------------------------
    ws = wb.create_sheet("Data")
    ws.append(["FiscalWeek", "Venue", "BlockKey", "Metric", "Value"])
    for row in data_rows:
        ws.append(list(row))
    ws.sheet_state = "hidden"

    # --- BlockLabels sheet -------------------------------------------------
    wsb = wb.create_sheet("BlockLabels")
    wsb.append(["FiscalWeek", "BlockKey", "DisplayLabel", "HelperKey"])
    for row in block_label_rows:
        wsb.append(list(row))
    wsb.sheet_state = "hidden"

    # --- Lists sheet (dropdown source) -------------------------------------
    wsl = wb.create_sheet("Lists")
    wsl.append(["FiscalWeek"])
    for week in weeks:
        wsl.append([week])
    wsl.sheet_state = "hidden"

    return len(data_rows), len(block_label_rows), len(weeks)


def write_dashboard(wb, weeks, venues, n_weeks_rows):
    ws = wb.create_sheet("Dashboard", 0)

    title_font = Font(name=FONT_NAME, size=14, bold=True)
    label_font = Font(name=FONT_NAME, size=10, bold=True)
    card_label_font = Font(name=FONT_NAME, size=9, color="FFFFFF")
    card_value_font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
    card_change_font = Font(name=FONT_NAME, size=9, italic=True, color="FFFFFF")
    header_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    subheader_font = Font(name=FONT_NAME, size=9, bold=True, color="FFFFFF")
    venue_font = Font(name=FONT_NAME, size=10)
    total_font = Font(name=FONT_NAME, size=10, bold=True)

    card_fill = PatternFill("solid", fgColor="2F5496")
    header_fill = PatternFill("solid", fgColor="1F3864")
    subheader_fill = PatternFill("solid", fgColor="4472C4")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    filter_fill = PatternFill("solid", fgColor="FFF2CC")

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Title + week filter ------------------------------------------------
    ws["A1"] = "Market Share Dashboard"
    ws["A1"].font = title_font

    ws["A2"] = "Fiscal Week:"
    ws["A2"].font = label_font
    default_week = DEFAULT_WEEK if DEFAULT_WEEK in weeks else (weeks[-1] if weeks else "")
    week_cell = ws["B2"]
    week_cell.value = default_week
    week_cell.font = Font(name=FONT_NAME, size=11, bold=True)
    week_cell.fill = filter_fill
    week_cell.alignment = Alignment(horizontal="center")
    week_cell.border = border

    dv = DataValidation(type="list", formula1=f"='Lists'!$A$2:$A${n_weeks_rows + 1}", allow_blank=False)
    dv.error = "Pick a fiscal week from the list."
    dv.errorTitle = "Invalid week"
    ws.add_data_validation(dv)
    dv.add(week_cell)

    # --- KPI cards (row 4-6), driven off the TOTAL venue in Data ------------
    def sumifs(venue, block, metric):
        return (f"SUMIFS(Data!$E:$E,Data!$A:$A,$B$2,Data!$B:$B,\"{venue}\","
                f"Data!$C:$C,\"{block}\",Data!$D:$D,\"{metric}\")")

    def pct_formula(curr_expr, prior_expr):
        return f"IF(OR({prior_expr}=0,{prior_expr}=\"\"),\"n/a\",({curr_expr}-{prior_expr})/{prior_expr}*100)"

    cards = [
        ("Total Weekend Revenue", sumifs(TOTAL_VENUE, "Weekend", "Gross"),
         sumifs(TOTAL_VENUE, "Weekend", "PY Gross"), "currency", "vs PY Weekend"),
        ("Total WTD Revenue", sumifs(TOTAL_VENUE, "WTD", "Gross"),
         sumifs(TOTAL_VENUE, "WTD", "PY Gross"), "currency", "vs PY WTD"),
        ("WTD Market Share", sumifs(TOTAL_VENUE, "WTD", "Share"),
         sumifs(TOTAL_VENUE, "WTD", "PY"), "percent_direct", "pts vs PY"),
        ("WTD Industry Gross", sumifs(TOTAL_VENUE, "WTD", "Industry"),
         sumifs(TOTAL_VENUE, "WTD", "PY Industry"), "currency", "vs PY Industry"),
    ]

    card_col = 1
    for label, value_expr, prior_expr, fmt, sublabel in cards:
        cs = get_column_letter(card_col)
        ce = get_column_letter(card_col + 1)
        for rr in (4, 5, 6):
            ws.merge_cells(f"{cs}{rr}:{ce}{rr}")

        lc = ws[f"{cs}4"]
        lc.value = label
        lc.font = card_label_font
        lc.fill = card_fill
        lc.alignment = Alignment(horizontal="center")

        vc = ws[f"{cs}5"]
        vc.value = f"={value_expr}"
        vc.font = card_value_font
        vc.fill = card_fill
        vc.alignment = Alignment(horizontal="center")
        vc.number_format = CURRENCY_FMT if fmt == "currency" else PERCENT_FMT

        cc = ws[f"{cs}6"]
        if fmt == "percent_direct":
            cc.value = f"=IFERROR(TEXT({value_expr}-{prior_expr},\"+0.0;-0.0\")&\" {sublabel}\",\"n/a {sublabel}\")"
        else:
            pct_expr = pct_formula(value_expr, prior_expr)
            cc.value = f"=IFERROR(IF({pct_expr}=\"n/a\",\"n/a\",TEXT({pct_expr},\"+0.0;-0.0\")&\"% {sublabel}\"),\"n/a {sublabel}\")"
        cc.font = card_change_font
        cc.fill = card_fill
        cc.alignment = Alignment(horizontal="center")

        card_col += 2

    # --- Main table ----------------------------------------------------------
    table_start_row = 8
    blockkey_row = table_start_row          # hidden helper row: raw block key
    block_header_row = table_start_row + 1  # visible: Weekend / WTD / actual date
    metric_header_row = table_start_row + 2
    first_data_row = table_start_row + 3

    ws.cell(row=block_header_row, column=1, value="Venue").font = header_font
    ws.cell(row=block_header_row, column=1).fill = header_fill
    ws.cell(row=metric_header_row, column=1).fill = header_fill
    ws.merge_cells(start_row=block_header_row, start_column=1,
                    end_row=metric_header_row, end_column=1)

    col = 2
    block_first_col = {}
    for block_key in BLOCK_KEYS:
        start_col = col
        end_col = col + len(METRICS) - 1
        block_first_col[block_key] = start_col

        # hidden helper row: the raw block key, repeated across this block's columns
        for i in range(len(METRICS)):
            hc = ws.cell(row=blockkey_row, column=start_col + i, value=block_key)
            hc.font = Font(name=FONT_NAME, size=1, color="FFFFFF")

        ws.merge_cells(start_row=block_header_row, start_column=start_col,
                        end_row=block_header_row, end_column=end_col)
        bh = ws.cell(row=block_header_row, column=start_col)
        if block_key in ("Weekend", "WTD"):
            bh.value = block_key
        else:
            helper_key_expr = f"$B$2&\"|\"&{get_column_letter(start_col)}${blockkey_row}"
            bh.value = (f"=IFERROR(INDEX(BlockLabels!$C:$C,"
                        f"MATCH({helper_key_expr},BlockLabels!$D:$D,0)),\"\")")
        bh.font = header_font
        bh.fill = header_fill
        bh.alignment = Alignment(horizontal="center")

        for i, metric in enumerate(METRICS):
            mh = ws.cell(row=metric_header_row, column=start_col + i, value=metric)
            mh.font = subheader_font
            mh.fill = subheader_fill
            mh.alignment = Alignment(horizontal="center")
            mh.border = border
        col = end_col + 1

    last_col = col - 1

    def venue_row_formula(venue_cell_ref, block_key_cell_ref, metric_literal):
        exists = (f"COUNTIFS(Data!$A:$A,$B$2,Data!$B:$B,{venue_cell_ref},"
                  f"Data!$C:$C,{block_key_cell_ref})")
        val = (f"SUMIFS(Data!$E:$E,Data!$A:$A,$B$2,Data!$B:$B,{venue_cell_ref},"
               f"Data!$C:$C,{block_key_cell_ref},Data!$D:$D,\"{metric_literal}\")")
        return f'=IF({exists}=0,"",{val})'

    r = first_data_row
    for v in venues:
        ws.cell(row=r, column=1, value=v).font = venue_font
        ws.cell(row=r, column=1).border = border
        venue_ref = f"$A{r}"
        c = 2
        for block_key in BLOCK_KEYS:
            block_key_ref = f"{get_column_letter(block_first_col[block_key])}${blockkey_row}"
            for metric in METRICS:
                cell = ws.cell(row=r, column=c, value=venue_row_formula(venue_ref, block_key_ref, metric))
                cell.font = venue_font
                cell.number_format = _fmt_for(metric)
                cell.border = border
                c += 1
        r += 1

    # TOTAL row
    ws.cell(row=r, column=1, value=TOTAL_VENUE).font = total_font
    ws.cell(row=r, column=1).fill = total_fill
    ws.cell(row=r, column=1).border = border
    venue_ref = f'"{TOTAL_VENUE}"'
    c = 2
    for block_key in BLOCK_KEYS:
        block_key_ref = f"{get_column_letter(block_first_col[block_key])}${blockkey_row}"
        for metric in METRICS:
            cell = ws.cell(row=r, column=c, value=venue_row_formula(venue_ref, block_key_ref, metric))
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = _fmt_for(metric)
            cell.border = border
            c += 1

    # --- Cosmetics -------------------------------------------------------
    ws.row_dimensions[blockkey_row].hidden = True
    ws.freeze_panes = ws.cell(row=first_data_row, column=2)
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13


def main():
    weeks, venues, data_rows, block_label_rows = collect_all_weeks()
    if not weeks:
        print("No fiscal weeks found in fct_wtd_daily_combined.")
        con.close()
        return

    wb = Workbook()
    wb.remove(wb.active)

    n_data_rows, n_label_rows, n_weeks_rows = write_hidden_sheets(wb, weeks, data_rows, block_label_rows)
    write_dashboard(wb, weeks, venues, n_weeks_rows)

    wb.save(OUTPUT_FILE)
    con.close()
    print(f"Saved {OUTPUT_FILE} — {len(weeks)} fiscal weeks, {len(venues)} venues.")
    print("Open the file and use the 'Fiscal Week' dropdown on the Dashboard sheet to filter.")


if __name__ == "__main__":
    main()
